from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path
from tkinter import filedialog

import customtkinter as ctk
from gui import theme
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from core.database import get_all_classes, get_attendance_by_date, log_export, verify_admin
from gui.widgets import ThemedDropdown, ThemedRangePicker, make_eye_image, center_dialog

COLLEGE_NAME = "Nihareeka College of Management and Information Technology"
STATUS_HEX = {
    "present": "#DFF5E1",
    "late": "#FFF4CC",
    "absent": "#FDE2E2",
}
STATUS_TEXT_HEX = {
    "present": "#2E7D32",
    "late": "#B28704",
    "absent": "#C62828",
}


class ExportPage(ctk.CTkFrame):
    def __init__(self, master, username: str, role: str) -> None:
        super().__init__(master, fg_color="transparent")
        self.master_frame = master
        self.username = username
        self.role = role
        self._classes: list[dict] = []
        self._class_map: dict[str, dict] = {}

        self.grid_columnconfigure(0, weight=1)
        # Centered container using grid with side padding
        container = ctk.CTkFrame(self, fg_color="transparent")
        container.grid(row=0, column=0, sticky="nsew", padx=60, pady=0)
        container.grid_columnconfigure(0, weight=1)

        main_card = ctk.CTkFrame(container, fg_color=theme.BG_SURFACE, corner_radius=12)
        main_card.grid(row=0, column=0, sticky="ew", padx=0, pady=(20, 12))
        main_card.grid_columnconfigure(0, weight=1)

        # Title inside main card
        ctk.CTkLabel(main_card, text="Export Reports", font=ctk.CTkFont(size=20, weight="bold"), text_color=theme.TEXT_PRIMARY).grid(row=0, column=0, sticky="w", padx=20, pady=(20, 6))
        ctk.CTkLabel(main_card, text="Set filters and choose an export format.", font=ctk.CTkFont(size=13), text_color=theme.TEXT_SECONDARY).grid(row=1, column=0, sticky="w", padx=20, pady=(0, 12))

        body = ctk.CTkFrame(main_card, fg_color="transparent")
        body.grid(row=2, column=0, sticky="nsew", padx=0, pady=(0, 16))
        body.grid_columnconfigure(0, weight=1)

        self._build_filters(body)
        self._build_actions(body)

        self.status_label = ctk.CTkLabel(
            body,
            text="",
            font=ctk.CTkFont(size=13),
            text_color=theme.TEXT_SECONDARY,
            justify="left",
            wraplength=760,
        )
        self.status_label.grid(row=3, column=0, padx=24, pady=(0, 20), sticky="w")

        self.refresh()

    def _build_filters(self, parent) -> None:
        filter_card = ctk.CTkFrame(parent, fg_color=theme.BG_SURFACE, corner_radius=12)
        filter_card.grid(row=0, column=0, sticky="ew", padx=20, pady=(0, 12))
        filter_card.grid_columnconfigure(1, weight=1)
        filter_card.grid_columnconfigure(3, weight=2)

        # Row 0 — Class selector
        ctk.CTkLabel(
            filter_card, text="Class",
            font=ctk.CTkFont(size=13), text_color=theme.TEXT_PRIMARY,
        ).grid(row=0, column=0, sticky="w", padx=(18, 8), pady=(16, 8))
        self.class_dropdown = ThemedDropdown(filter_card, values=["Loading…"], height=36)
        self.class_dropdown.grid(row=0, column=1, sticky="ew", padx=(0, 24), pady=(16, 8))

        # Row 0 — Date range picker
        ctk.CTkLabel(
            filter_card, text="Date Range",
            font=ctk.CTkFont(size=13), text_color=theme.TEXT_PRIMARY,
        ).grid(row=0, column=2, sticky="w", padx=(0, 8), pady=(16, 8))
        self.range_picker = ThemedRangePicker(
            filter_card,
            initial_from=date.today(),
            initial_to=date.today(),
            height=36,
        )
        self.range_picker.grid(row=0, column=3, sticky="ew", padx=(0, 18), pady=(10, 8))

    def _build_actions(self, parent) -> None:
        action_card = ctk.CTkFrame(parent, fg_color=theme.BG_SURFACE, corner_radius=12)
        action_card.grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 16))
        # Centered action buttons with fixed widths, placed side-by-side
        action_card.grid_columnconfigure(0, weight=1)
        action_card.grid_columnconfigure(1, weight=1)
        btn_frame = ctk.CTkFrame(action_card, fg_color="transparent")
        btn_frame.grid(row=0, column=0, columnspan=2, pady=16)
        btn_frame.grid_columnconfigure(0, weight=1)
        btn_frame.grid_columnconfigure(1, weight=1)

        pdf_btn = ctk.CTkButton(
            btn_frame,
            text="Export PDF",
            command=self.handle_export_pdf,
            fg_color=theme.ACCENT,
            hover_color=theme.ACCENT_HOVER,
            height=36,
            width=160,
            font=ctk.CTkFont(size=14, weight="bold"),
        )
        excel_btn = ctk.CTkButton(
            btn_frame,
            text="Export to Excel",
            command=self.handle_export_excel,
            fg_color=theme.BTN_SUCCESS,
            hover_color=theme.BTN_SUCCESS_HVR,
            height=36,
            width=160,
            font=ctk.CTkFont(size=14, weight="bold"),
        )
        try:
            pdf_btn.pack(side="left", padx=(6, 12))
            excel_btn.pack(side="left", padx=(12, 6))
        except Exception:
            pdf_btn.grid(row=0, column=0, padx=(6, 12))
            excel_btn.grid(row=0, column=1, padx=(12, 6))

    def refresh(self) -> None:
        try:
            self._classes = get_all_classes()
        except Exception as error:
            self._set_status(f"Failed to load classes: {error}", "error")
            return

        current_value = self.class_dropdown.get().strip() if hasattr(self, "class_dropdown") else ""
        self._class_map.clear()
        options = []
        for class_row in self._classes:
            label = self._class_label(class_row)
            self._class_map[label] = class_row
            options.append(label)

        if not options:
            options = ["No classes available"]

        self.class_dropdown.configure(values=options)
        self.class_dropdown.set(current_value if current_value in options else options[0])

    def _class_label(self, class_row: dict) -> str:
        return f"{class_row.get('name', '')} - {class_row.get('section', '')}"

    def _selected_class(self) -> dict | None:
        return self._class_map.get(self.class_dropdown.get().strip())

    def _date_range(self) -> tuple[date, date]:
        from_date = self.range_picker.get_from_date()
        to_date   = self.range_picker.get_to_date()
        if not from_date or not to_date:
            raise ValueError("Please select both a start and end date.")
        if from_date > to_date:
            raise ValueError("From date cannot be later than To date.")
        return from_date, to_date

    def _collect_rows(self, from_date: date, to_date: date, class_id: int | None) -> list[dict]:
        rows: list[dict] = []
        current = from_date
        while current <= to_date:
            rows.extend(get_attendance_by_date(current, class_id))
            current += timedelta(days=1)
        rows.sort(key=lambda item: (str(item.get("date", "")), str(item.get("time", "")), str(item.get("student_id", ""))))
        return rows

    def _password_dialog(self) -> str | None:
        dialog = ctk.CTkToplevel(self)
        dialog.title("Confirm Export")
        dialog.resizable(False, False)
        dialog.transient(self.winfo_toplevel())
        dialog.grab_set()
        center_dialog(dialog, 420, 248)

        dialog.grid_columnconfigure(0, weight=1)
        dialog.grid_rowconfigure(0, weight=1)

        result = {"password": None}

        card = ctk.CTkFrame(dialog, fg_color=theme.BG_SURFACE, corner_radius=0)
        card.grid(row=0, column=0, sticky="nsew")
        card.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            card, text="Export Confirmation",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color=theme.TEXT_PRIMARY,
        ).grid(row=0, column=0, sticky="w", padx=20, pady=(20, 6))

        ctk.CTkLabel(
            card,
            text="Enter your admin password to continue with this export.",
            justify="left",
            wraplength=360,
            font=ctk.CTkFont(size=13),
            text_color=theme.TEXT_SECONDARY,
        ).grid(row=1, column=0, sticky="w", padx=20, pady=(0, 12))

        # Password row: entry + eye toggle (same pattern as Login page)
        pw_container = ctk.CTkFrame(card, fg_color=theme.BG_SURFACE_ALT, corner_radius=6)
        pw_container.grid(row=2, column=0, sticky="ew", padx=20, pady=(0, 18))
        pw_container.grid_columnconfigure(0, weight=1)

        password_entry = ctk.CTkEntry(
            pw_container,
            show="*",
            placeholder_text="Admin password",
            height=42,
            fg_color=theme.BG_SURFACE_ALT,
            border_width=0,
            text_color=theme.TEXT_PRIMARY,
            placeholder_text_color=theme.TEXT_MUTED,
            corner_radius=0,
        )
        password_entry.grid(row=0, column=0, sticky="ew")
        password_entry._entry.focus_set()

        _eye_pil       = make_eye_image(18, theme.TEXT_MUTED, slashed=False)
        _eye_slash_pil = make_eye_image(18, theme.TEXT_MUTED, slashed=True)
        if _eye_pil is not None:
            _icon_eye       = ctk.CTkImage(light_image=_eye_pil,       dark_image=_eye_pil,       size=(18, 18))
            _icon_eye_slash = ctk.CTkImage(light_image=_eye_slash_pil, dark_image=_eye_slash_pil, size=(18, 18))
            _eye_kw: dict = {"text": "", "image": _icon_eye_slash}
        else:
            _icon_eye = _icon_eye_slash = None
            _eye_kw = {"text": "👁"}

        def _toggle_pw() -> None:
            if password_entry.cget("show") == "":
                password_entry.configure(show="*")
                eye_btn.configure(**({"image": _icon_eye_slash, "text": ""} if _icon_eye_slash else {"text": "👁"}))
            else:
                password_entry.configure(show="")
                eye_btn.configure(**({"image": _icon_eye, "text": ""} if _icon_eye else {"text": "🙈"}))

        eye_btn = ctk.CTkButton(
            pw_container, width=34, height=34,
            fg_color="transparent", hover_color=theme.INPUT_HIGHLIGHT,
            text_color=theme.TEXT_MUTED, corner_radius=4,
            command=_toggle_pw, **_eye_kw,
        )
        eye_btn.grid(row=0, column=1, padx=(0, 4))

        def submit() -> None:
            result["password"] = password_entry.get()
            dialog.destroy()

        actions = ctk.CTkFrame(card, fg_color="transparent")
        actions.grid(row=3, column=0, sticky="ew", padx=20, pady=(0, 20))

        ctk.CTkButton(
            actions, text="Cancel", command=dialog.destroy,
            fg_color=theme.BG_SURFACE_ALT, hover_color=theme.BG_HOVER,
            text_color=theme.TEXT_SECONDARY, corner_radius=6,
        ).pack(side="left")
        ctk.CTkButton(
            actions, text="Confirm", command=submit,
            fg_color=theme.ACCENT, hover_color=theme.ACCENT_HOVER,
            text_color=theme.TEXT_PRIMARY, corner_radius=6,
        ).pack(side="right")

        password_entry.bind("<Return>", lambda _: submit())
        self.wait_window(dialog)
        return result["password"]

    def _verify_export_password(self) -> bool:
        admin_username = self.username
        if not admin_username:
            self._notify("Invalid credentials. Export cancelled.", "error")
            self._set_status("Export cancelled because no logged-in admin was found.", "error")
            return False

        password = self._password_dialog()
        if not password or verify_admin(admin_username, password) is None:
            self._notify("Invalid credentials. Export cancelled.", "error")
            self._set_status("Invalid credentials. Export cancelled.", "error")
            return False
        return True

    def _report_metadata(self) -> tuple[date, date, dict, int, str]:
        from_date, to_date = self._date_range()
        selected_class = self._selected_class()
        if selected_class is None:
            raise ValueError("Please select a class before exporting.")
        class_id   = int(selected_class["id"])
        class_name = f"{selected_class.get('name', '')} {selected_class.get('section', '')}".strip()
        return from_date, to_date, selected_class, class_id, class_name

    def _build_export_dataframe(self, rows: list[dict]) -> pd.DataFrame:
        data_rows = []
        for row in rows:
            status = str(row.get("status", "")).lower()
            data_rows.append(
                {
                    "Student ID": row.get("student_id", ""),
                    "Name": row.get("name", ""),
                    "Class": self._class_name_from_id(row.get("class_id")),
                    "Date": str(row.get("date", "")),
                    "Time": str(row.get("time", ""))[:8],
                    "Status": status.capitalize() if status else "",
                }
            )
        return pd.DataFrame(data_rows, columns=["Student ID", "Name", "Class", "Date", "Time", "Status"])

    def _class_name_from_id(self, class_id: int | None) -> str:
        if class_id is None:
            return ""
        for class_row in self._classes:
            if int(class_row.get("id", -1)) == int(class_id):
                return f"{class_row.get('name', '')} {class_row.get('section', '')}".strip()
        return str(class_id)

    def _export_pdf(self, report_path: Path, rows: list[dict], from_date: date, to_date: date, class_name: str) -> Path:
        styles = getSampleStyleSheet()
        story: list = []

        story.append(Paragraph(COLLEGE_NAME, styles["Title"]))
        story.append(Spacer(1, 8))
        story.append(Paragraph("Attendance Export Report", styles["Heading2"]))
        story.append(Spacer(1, 8))
        story.append(Paragraph(f"Date Range: {from_date} to {to_date}", styles["Normal"]))
        story.append(Paragraph(f"Class: {class_name}", styles["Normal"]))
        story.append(Spacer(1, 14))

        table_data = [["Student ID", "Name", "Class", "Date", "Time", "Status"]]
        for row in rows:
            status = str(row.get("status", "")).lower()
            table_data.append(
                [
                    str(row.get("student_id", "")),
                    str(row.get("name", "")),
                    self._class_name_from_id(row.get("class_id")),
                    str(row.get("date", "")),
                    str(row.get("time", ""))[:8],
                    status.capitalize() if status else "",
                ]
            )
        if len(table_data) == 1:
            table_data.append(["No records found", "", "", "", "", ""])

        table = Table(table_data, colWidths=[85, 130, 110, 75, 60, 70])
        style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E88E5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F8F8F8"), colors.white]),
        ]
        for row_index, row in enumerate(rows, start=1):
            status = str(row.get("status", "")).lower()
            fill_hex = STATUS_HEX.get(status)
            text_hex = STATUS_TEXT_HEX.get(status)
            if fill_hex and text_hex:
                style_commands.append(("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor(fill_hex)))
                style_commands.append(("TEXTCOLOR", (5, row_index), (5, row_index), colors.HexColor(text_hex)))
                style_commands.append(("FONTNAME", (5, row_index), (5, row_index), "Helvetica-Bold"))
        table.setStyle(TableStyle(style_commands))
        story.append(table)

        document = SimpleDocTemplate(
            str(report_path),
            pagesize=landscape(A4),
            leftMargin=28,
            rightMargin=28,
            topMargin=28,
            bottomMargin=28,
        )
        document.build(story)
        return report_path

    def _export_excel(self, report_path: Path, dataframe: pd.DataFrame) -> Path:
        with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
            dataframe.to_excel(writer, index=False, sheet_name="Attendance")
            worksheet = writer.sheets["Attendance"]

            header_fill = PatternFill(fill_type="solid", fgColor="1E88E5")
            header_font = Font(color="FFFFFF", bold=True)
            center = Alignment(horizontal="center", vertical="center")
            status_fills = {
                "Present": PatternFill(fill_type="solid", fgColor="DFF5E1"),
                "Late": PatternFill(fill_type="solid", fgColor="FFF4CC"),
                "Absent": PatternFill(fill_type="solid", fgColor="FDE2E2"),
            }
            status_fonts = {
                "Present": Font(color="2E7D32", bold=True),
                "Late": Font(color="B28704", bold=True),
                "Absent": Font(color="C62828", bold=True),
            }

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center

            widths = {"A": 18, "B": 28, "C": 22, "D": 14, "E": 12, "F": 14}
            for column_name, width in widths.items():
                worksheet.column_dimensions[column_name].width = width

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = center
                status_cell = row[5]
                if status_cell.value in status_fills:
                    status_cell.fill = status_fills[status_cell.value]
                    status_cell.font = status_fonts[status_cell.value]

        return report_path

    def _export_common(self, export_type: str) -> None:
        try:
            from_date, to_date, _selected_class, class_id, class_name = self._report_metadata()
        except Exception as error:
            self._set_status(str(error), "error")
            self._notify(str(error), "error")
            return

        if not self._verify_export_password():
            return

        try:
            rows = self._collect_rows(from_date, to_date, class_id)
        except Exception as error:
            self._set_status(f"Failed to collect attendance records: {error}", "error")
            self._notify("Export failed.", "error")
            return

        export_dir = Path(__file__).resolve().parent.parent / "attendance"
        export_dir.mkdir(parents=True, exist_ok=True)
        date_suffix = f"{from_date}_to_{to_date}"

        if export_type == "pdf":
            selected_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                initialdir=str(export_dir),
                initialfile=f"attendance_report_{date_suffix}.pdf",
                title="Save PDF Report",
            )
        else:
            selected_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialdir=str(export_dir),
                initialfile=f"attendance_report_{date_suffix}.xlsx",
                title="Save Excel Report",
            )

        if not selected_path:
            self._set_status("Export cancelled.", "error")
            self._notify("Export cancelled.", "error")
            return

        report_path = Path(selected_path)
        try:
            if export_type == "pdf":
                self._export_pdf(report_path, rows, from_date, to_date, class_name)
            else:
                dataframe = self._build_export_dataframe(rows)
                self._export_excel(report_path, dataframe)

            log_export(self.username or "system", export_type, report_path.name)
        except Exception as error:
            self._set_status(f"Export failed: {error}", "error")
            self._notify("Export failed.", "error")
            return

        self._set_status(f"{export_type.upper()} report saved to {report_path}", "success")
        self._notify("Export successful.", "success")

    def handle_export_pdf(self) -> None:
        self._export_common("pdf")

    def handle_export_excel(self) -> None:
        self._export_common("excel")

    def _set_status(self, message: str, kind: str) -> None:
        self.status_label.configure(text=message, text_color="#6FD36F" if kind == "success" else "#FF5252")

    def _notify(self, message: str, kind: str) -> None:
        try:
            self.winfo_toplevel().show_notification(message, kind)
        except Exception:
            pass
